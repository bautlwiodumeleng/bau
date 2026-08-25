from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, abort, current_app, send_from_directory
import csv
from io import StringIO
import requests

from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import login_user, logout_user, login_required, current_user
from openpyxl import Workbook

from app.forms import ResetPasswordForm
from .forms import ForgotPasswordForm, RegisterForm, LoginForm, CustomerForm, SearchForm, ProfileForm, TaskForm
from .models import User, Customer, Task, CustomerDocument
from . import db
from flask import abort
from datetime import datetime, UTC, date, timedelta

from openpyxl.styles import PatternFill

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
import cloudinary
import cloudinary.uploader
from .cloudinary_config import configure_cloudinary
from .models import User, Customer, Task, CustomerDocument
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Border, Side
from datetime import datetime
import secrets
from datetime import datetime, timedelta

from reportlab.platypus.flowables import HRFlowable
from flask import Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    HRFlowable
)

main = Blueprint("main", __name__)
    
@main.route("/")
def home():
    return render_template("home.html")

@main.route("/portal-information")
def portal_information():
    return render_template("portal_information.html")

@main.route("/register", methods=["GET", "POST"])
def register():

    form = RegisterForm()

    if form.validate_on_submit():

        existing_user = User.query.filter_by(username=form.username.data).first()

        if existing_user:
            flash("Username already exists.", "danger")
            return redirect(url_for("main.register"))
        existing_email = User.query.filter_by(email=form.email.data).first()

        if existing_email:
            flash("An account with this email already exists.", "danger")
            return redirect(url_for("main.register"))    

        user = User(
            username=form.username.data,
            email=form.email.data,
            password=generate_password_hash(form.password.data)
        )

        db.session.add(user)
        db.session.commit()

        flash("Registration successful! You can now log in.", "success")

        return redirect(url_for("main.home"))

    return render_template("register.html", form=form)

@main.route("/login", methods=["GET", "POST"])
def login():

    form = LoginForm()

    if form.validate_on_submit():

        user = User.query.filter_by(username=form.username.data).first()

        if user and check_password_hash(user.password, form.password.data):

            login_user(user, remember=form.remember.data)

            flash("Logged in successfully!", "success")

            return redirect(url_for("main.home"))

        flash("Invalid username or password. Please try again.", "danger")

    return render_template("login.html", form=form)


@main.route("/dashboard")
@login_required
def dashboard():

    # Total customers belonging to the logged-in user
    total_customers = Customer.query.filter_by(
        created_by=current_user.id
    ).count()

    # Total organisations belonging to the logged-in user's customers
    total_organisations = (
        db.session.query(Customer.organisation_name)
        .filter(
            Customer.created_by == current_user.id,
            Customer.organisation_name.isnot(None),
            Customer.organisation_name != ""
        )
        .distinct()
        .count()
    )

    now = datetime.now(UTC)

    # Customers added this month by the logged-in user
    new_this_month = Customer.query.filter(
        Customer.created_by == current_user.id,
        Customer.created_at >= datetime(
            now.year,
            now.month,
            1,
            tzinfo=UTC
        )
    ).count()

    # Five most recent customers belonging to the logged-in user
    recent_customers = (
        Customer.query
        .filter_by(created_by=current_user.id)
        .order_by(Customer.id.desc())
        .limit(5)
        .all()
    )

    today = date.today()

    # Upcoming tasks belonging to the logged-in user's customers
    upcoming_tasks = (
        Task.query
        .join(Customer)
        .filter(
            Customer.created_by == current_user.id,
            Task.completed == False,
            Task.due_date != None,
            Task.due_date >= today
        )
        .order_by(Task.due_date)
        .limit(5)
        .all()
    )

    # Overdue tasks belonging to the logged-in user's customers
    overdue_tasks = (
        Task.query
        .join(Customer)
        .filter(
            Customer.created_by == current_user.id,
            Task.completed == False,
            Task.due_date != None,
            Task.due_date < today
        )
        .order_by(Task.due_date)
        .all()
    )

    # Total tasks belonging to the logged-in user's customers
    total_tasks = (
        Task.query
        .join(Customer)
        .filter(
            Customer.created_by == current_user.id
        )
        .count()
    )

    return render_template(
        "dashboard.html",
        total_customers=total_customers,
        total_organisations=total_organisations,
        total_tasks=total_tasks,
        new_this_month=new_this_month,
        recent_customers=recent_customers,
        upcoming_tasks=upcoming_tasks,
        overdue_tasks=overdue_tasks
    )


    
@main.route("/logout")
@login_required
def logout():

    logout_user()

    flash("You have been logged out.", "info")

    return redirect(url_for("main.home"))


@main.route("/add_customer", methods=["GET", "POST"])
@login_required
def add_customer():

    form = CustomerForm()

    if form.validate_on_submit():

        existing_customer = Customer.query.filter(
            db.func.lower(Customer.name) == form.name.data.lower(),
            Customer.created_by == current_user.id
        ).first()

        if existing_customer:
            flash("Customer already exists.", "danger")
            return redirect(url_for("main.add_customer"))

        # Find the highest existing customer ID
        highest_id = db.session.query(
            db.func.max(Customer.id)
        ).scalar()

        next_id = (highest_id or 0) + 1

        # Generate the customer code before inserting
        customer_code = f"CUST-{next_id:06d}"

        customer = Customer(
            customer_code=customer_code,
            name=form.name.data,
            phone=form.phone.data,
            email=form.email.data,
            organisation_name=form.organisation_name.data,
            address=form.address.data,
            notes=form.notes.data,
            created_by=current_user.id
        )

        db.session.add(customer)
        db.session.commit()

        flash("Customer added successfully!", "success")

        return redirect(url_for("main.customers"))

    return render_template(
        "add_customer.html",
        form=form
    )






# ⬇️ PUT THE NEW ROUTE HERE


@main.route("/add_task/<string:customer_code>", methods=["GET", "POST"])
@login_required
def add_task(customer_code):

    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id
    ).first_or_404()

    form = TaskForm()

    if form.validate_on_submit():

        task = Task(
            title=form.title.data,
            due_date=form.due_date.data,
            priority=form.priority.data,
            customer_id=customer.id
        )

        db.session.add(task)
        db.session.commit()

        flash("Task added successfully!", "success")

        return redirect(
            url_for(
                "main.customer_details",
                customer_code=customer.customer_code
            )
        )

    return render_template(
        "add_task.html",
        form=form,
        customer=customer
    )

@main.route("/complete_task/<int:task_id>")
@login_required
def complete_task(task_id):

    task = Task.query.get_or_404(task_id)

    if task.customer.created_by != current_user.id:
        abort(403)

    task.completed = True
    task.completed_at = datetime.now(UTC)

    db.session.commit()

    flash("Task completed!", "success")

    return redirect(
        url_for(
            "main.customer_details",
            customer_code=task.customer.customer_code
        )
    )

@main.route("/edit_task/<int:task_id>", methods=["GET", "POST"])
@login_required
def edit_task(task_id):

    task = Task.query.get_or_404(task_id)

    if task.customer.created_by != current_user.id:
        abort(403)

    form = TaskForm(obj=task)

    if form.validate_on_submit():

        task.title = form.title.data
        task.due_date = form.due_date.data
        task.priority = form.priority.data

        db.session.commit()

        flash("Task updated successfully!", "success")

        return redirect(
            url_for(
                "main.customer_details",
                customer_code=task.customer.customer_code
            )
        )

    return render_template(
        "edit_task.html",
        form=form,
        task=task
    )


    
@main.route("/delete_task/<int:task_id>")
@login_required
def delete_task(task_id):

    task = Task.query.get_or_404(task_id)

    if task.customer.created_by != current_user.id:
        abort(403)

    customer_code = task.customer.customer_code

    db.session.delete(task)
    db.session.commit()

    flash("Task deleted successfully!", "success")

    return redirect(
        url_for(
            "main.customer_details",
            customer_code=customer_code
        )
    )

# ⬇️ THEN YOUR NEXT ROUTE CONTINUES

@main.route("/customers")
@login_required
def customers():

    search = request.args.get("search", "").strip()

    query = Customer.query.filter_by(
        created_by=current_user.id,
        is_archived=False
    )

    if search:

        query = query.filter(
            (Customer.name.ilike(f"%{search}%")) |
            (Customer.phone.ilike(f"%{search}%")) |
            (Customer.email.ilike(f"%{search}%")) |
            (Customer.organisation_name.ilike(f"%{search}%"))
        )

    customer_list = query.all()

    print("Browser customer count:", len(customer_list))

    for c in customer_list:
        print(
            c.customer_code,
            c.organisation_name,
            c.name
        )

    return render_template(
        "customers.html",
        customers=customer_list
    )
    
@main.route("/customer/<string:customer_code>/edit", methods=["GET", "POST"])
@login_required
def edit_customer(customer_code):



    customer = Customer.query.filter_by( customer_code=customer_code, created_by=current_user.id ).first_or_404()

    form = CustomerForm(obj=customer)

    if form.validate_on_submit():

        customer.name = form.name.data
        customer.phone = form.phone.data
        customer.email = form.email.data
        customer.organisation_name = form.organisation_name.data
        customer.address = form.address.data
        customer.notes = form.notes.data

        db.session.commit()

        flash("Customer updated successfully!", "success")

        return redirect(url_for("main.customers"))

    return render_template(
        "edit_customer.html",
        form=form
    )
    
@main.route("/live_search")
@login_required
def live_search():

    query = request.args.get("q", "")

    customers = Customer.query.filter(
        Customer.created_by == current_user.id,
        Customer.name.ilike(f"%{query}%")
    ).all()

    return jsonify([
        {
            "code": customer.code,
            "name": customer.name,
            "phone": customer.phone,
            "email": customer.email,
            "organisation_name": customer.organisation_name,
            "address": customer.address
        }
        for customer in customers
    ])

@main.route("/export/csv")
@login_required
def export_csv():

    customers = Customer.query.filter_by(
        created_by=current_user.id,
        is_archived=False
    ).all()

    print("Export customer count:", len(customers))
    
    output = StringIO()
    writer = csv.writer(output)

    # Column headers
    writer.writerow([
        "Customer Code",
        "Name",
        "Phone",
        "Email",
        "Organisation",
        "Address",
        "Notes"
    ])

    # Customer data
    for customer in customers:

                writer.writerow([
                    format_customer_code(customer.customer_code),
                    customer.name,
                    customer.phone,
                    customer.email,
                    customer.organisation_name,
                    customer.address,
                    customer.notes
                ])   

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=customers.csv"
        }
    )
    
    
@main.route("/export/xlsx")
@login_required
def export_xlsx():

    customers = Customer.query.filter_by(
        created_by=current_user.id,
        is_archived=False
    ).all()
    
    print("Export customer count:", len(customers))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header row
    headers = [
        "Customer Code",
        "Name",
        "Phone",
        "Email",
        "Organisation",
        "Address",
        "Notes"
    ]

    ws.append(headers)

    # Header formatting
    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = header_fill

    # Customer data
    # Customer data
    for customer in customers:

        ws.append([
            format_customer_code(customer.customer_code),
            customer.name or "",
            customer.phone or "",
            customer.email or "",
            customer.organisation_name or "",
            customer.address or "",
            customer.notes or ""
        ])
            
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column_letter].width = max_length + 3

        
    # Freeze the header row
    ws.freeze_panes = "A2"

    # Thin borders for all cells
    thin = Side(
        border_style="thin",
        color="000000"
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return Excel file
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=customers.xlsx"
        }
    ) 
    
@main.route("/export/customers")
@login_required
def export_customers():
    return render_template("export_customers.html")

@main.route("/export/pdf")
@login_required
def export_pdf():

    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)

    query = Customer.query.filter_by(
        created_by=current_user.id,
        is_archived=False
    )

    if month and year:

        start_date = datetime(
            year,
            month,
            1,
            tzinfo=UTC
        )

        if month == 12:
            end_date = datetime(
                year + 1,
                1,
                1,
                tzinfo=UTC
            )
        else:
            end_date = datetime(
                year,
                month + 1,
                1,
                tzinfo=UTC
            )

        query = query.filter(
            Customer.created_at >= start_date,
            Customer.created_at < end_date
        )

    customers = query.all()

    print(
        "PDF Export customer count:",
        len(customers)
    )

    print("PDF Export customer count:", len(customers))

    customer_count = len(customers)

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=5,
        rightMargin=18,
        topMargin=5,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()

    # ----------------------------------------------------
    # Compact Professional Styles
    # ----------------------------------------------------

    organisation_style = ParagraphStyle(
    "organisation",
    parent=styles["Normal"],
    fontName="Helvetica",
    fontSize=10,
    leading=11,
    alignment=1,
    spaceAfter=1
    )

    title_style = ParagraphStyle(
        "title",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=15,
        alignment=1,
        spaceAfter=2
    )

    report_style = ParagraphStyle(
        "report",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        alignment=0,
        spaceAfter=2
    )

    small_style = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontSize=8,
        leading=9
    )

    elements = []

    # ----------------------------------------------------
    # Organisation Header
    # ----------------------------------------------------

    organisation_info = f"<b>{get_organisation_name()}</b><br/>"

    if current_user.business_address:
        organisation_info += f"{current_user.business_address}<br/>"

    if current_user.business_phone:
        organisation_info += f"Tel: {current_user.business_phone}<br/>"

    if current_user.business_email:
        organisation_info += f"Email: {current_user.business_email}<br/>"

    if current_user.website:
        organisation_info += f"Website: {current_user.website}<br/>"

    # ----------------------------------------------------
    # Organisation Logo
    # ----------------------------------------------------

    logo = None

    if current_user.organisation_logo:
        try:
            logo_response = requests.get(
                current_user.organisation_logo,
                timeout=10
            )

            logo_response.raise_for_status()

            logo_file = BytesIO(logo_response.content)

            logo = Image(
                logo_file,
                width=70,
                height=50
            )

        except Exception as e:
            print(
                "Could not load organisation logo:",
                e
            )

    # ----------------------------------------------------
    # Organisation Header Table
    # ----------------------------------------------------

    if logo:

        header_table = Table(
            [
                [
                    logo,
                    Paragraph(
                        organisation_info,
                        organisation_style
                    )
                ]
            ],
            colWidths=[90, 600]
        )

    else:

        header_table = Table(
            [
                [
                    Paragraph(
                        organisation_info,
                        organisation_style
                    )
                ]
            ],
            colWidths=[690]
        )
    header_table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "LEFT"),

            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])
    )

    elements.append(header_table)

    elements.append(Spacer(1, 3))

    # ----------------------------------------------------
    # Single Separator
    # ----------------------------------------------------

    elements.append(
        HRFlowable(
            width="100%",
            thickness=0.6,
            color=colors.grey
        )
    )

    elements.append(Spacer(1, 4))

    # ----------------------------------------------------
    # Report Title
    # ----------------------------------------------------
    if month and year:
        report_period = datetime(
            year,
            month,
            1
        ).strftime("%B %Y")

        report_title = (
            f"<b>CUSTOMER MANAGEMENT SYSTEM — "
            f"{report_period}</b>"
        )

    else:
        report_title = (
            "<b>CUSTOMER MANAGEMENT SYSTEM — "
            "All Customers</b>"
        )

    elements.append(
        Paragraph(
            report_title,
            title_style
        )
    )

    # ----------------------------------------------------
    # Report Information
    # ----------------------------------------------------

    generated = datetime.now().strftime("%d %b %Y %H:%M")

    elements.append(
        Paragraph(
            f"<b>Customer List</b>"
            f"&nbsp;&nbsp;&nbsp;&nbsp;"
            f"Generated: {generated}"
            f"&nbsp;&nbsp;&nbsp;&nbsp;"
            f"Active Customers: <b>{customer_count}</b>",
            report_style
        )
    )

    elements.append(Spacer(1, 6))

    # ----------------------------------------------------
    # Customer List Table
    # ----------------------------------------------------

    table_data = [
        [
            "Customer ID",
            "Customer Name",
            "Phone",
            "Email",
            "Organisation",
            "Address"
        ]
    ]

    for customer in customers:

        table_data.append(
            [
                customer.customer_code or "",
                customer.name or "",
                customer.phone or "",
                customer.email or "",
                customer.organisation_name or "",
                customer.address or ""
            ]
        )

    customer_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            75,
            120,
            80,
            120,
            100,
            120
        ]
    )

    customer_table.setStyle(
        TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),

            # Borders
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

            # Alignment
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            # Padding
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
    )

    # ----------------------------------------------------
    # Zebra Striping
    # ----------------------------------------------------

    for row in range(1, len(table_data)):

        if row % 2 == 0:
            organisation_style = ParagraphStyle(
                "organisation",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=10,
                leading=11,
                alignment=1,
                spaceAfter=1
            )
            customer_table.setStyle(
                TableStyle([
                    (
                        "BACKGROUND",
                        (0, row),
                        (-1, row),
                        colors.whitesmoke
                    )
                ])
            )

    elements.append(customer_table)

    # ----------------------------------------------------
    # Build PDF
    # ----------------------------------------------------

    doc.build(
        elements,
        onFirstPage=pdf_footer,
        onLaterPages=pdf_footer
    )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="customer_report.pdf",
        mimetype="application/pdf"
    )


# --------------------------------------------------------
# Helper Functions
# --------------------------------------------------------

def format_customer_code(customer_code):
    return f"{customer_code}"


def get_organisation_name():
    return (
        current_user.organisation_name
        or "BAU Technologies (Pty) Ltd"
    )


def pdf_footer(canvas, doc):

    canvas.saveState()

    width, height = landscape(A4)

    canvas.setFont("Helvetica", 8)

    # Footer text
    canvas.drawString(
        20,
        15,
        "Powered by BAU Technologies (Pty) Ltd"
    )

    # Page number
    canvas.drawRightString(
        width - 20,
        15,
        f"Page {doc.page}"
    )

    canvas.restoreState()
    
    


# --------------------------------------------------------
# Customer Details
# --------------------------------------------------------

@main.route("/customer/<string:customer_code>")
@login_required
def customer_details(customer_code):

    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id
    ).first_or_404()

    return render_template(
        "customer_details.html",
        customer=customer
    )

@main.route(
    "/customer/<string:customer_code>/upload-document",
    methods=["POST"]
    )
@login_required
def upload_customer_document(customer_code):

    # Find the customer and make sure it belongs to the logged-in user
    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id
    ).first_or_404()

    # Check that a file was actually submitted
    if "document" not in request.files:
        flash("No file was selected.", "danger")
        return redirect(
            url_for(
                "main.customer_details",
                customer_code=customer.customer_code
            )
        )

    file = request.files["document"]

    # Check that the filename is not empty
    if file.filename == "":
        flash("No file was selected.", "danger")
        return redirect(
            url_for(
                "main.customer_details",
                customer_code=customer.customer_code
            )
        )

    # Allowed file types
    allowed_extensions = {
        "pdf",
        "doc",
        "docx",
        "xls",
        "xlsx",
        "jpg",
        "jpeg",
        "png"
    }

    filename = file.filename

    extension = (
        filename.rsplit(".", 1)[-1].lower()
        if "." in filename
        else ""
    )

    if extension not in allowed_extensions:
        flash(
            "File type not allowed. Please upload PDF, Word, Excel, JPG or PNG files.",
            "danger"
        )
        return redirect(
            url_for(
                "main.customer_details",
                customer_code=customer.customer_code
            )
        )

    # Generate a secure unique filename
    original_filename = secure_filename(filename)

    stored_filename = (
        f"{customer.id}_{uuid.uuid4().hex}_{original_filename}"
    )

    # Configure Cloudinary
    configure_cloudinary()

    # Upload the customer document to Cloudinary
    upload_result = cloudinary.uploader.upload(
        file,
        folder=f"customer_documents/{customer.customer_code}",
        resource_type="auto"
    )

    cloudinary_public_id = upload_result["public_id"]
    cloudinary_resource_type = upload_result["resource_type"]

    # Create database record
    document = CustomerDocument(
        customer_id=customer.id,
        filename=original_filename,
        stored_filename=stored_filename,
        cloudinary_public_id=cloudinary_public_id,
        cloudinary_resource_type=cloudinary_resource_type,
        file_type=file.content_type
    )

    db.session.add(document)
    db.session.commit()

    flash(
        "Document uploaded successfully.",
        "success"
    )

    return redirect(
        url_for(
            "main.customer_details",
            customer_code=customer.customer_code
        )
    )

@login_required
def customer_document(document_id, customer_code, action):

    # Make sure the customer belongs to the logged-in user
    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id
    ).first_or_404()

    # Find the document belonging to this customer
    document = CustomerDocument.query.filter_by(
        id=document_id,
        customer_id=customer.id
    ).first_or_404()

    # Only allow these two actions
    if action not in ["view", "download"]:
        abort(404)

    # Make sure this document has a Cloudinary file
    if not document.cloudinary_public_id:
        flash(
            "This document is not stored in Cloudinary.",
            "danger"
        )
        return redirect(
            url_for(
                "main.customer_details",
                customer_code=customer.customer_code
            )
        )


    # Configure Cloudinary
    configure_cloudinary()

    resource_type = document.cloudinary_resource_type or "image"

    # Generate the Cloudinary URL
    url, options = cloudinary.utils.cloudinary_url(
        document.cloudinary_public_id,
        resource_type=resource_type,
        secure=True
    )

    if action == "view":
        return redirect(url)

    # Download
    download_url, options = cloudinary.utils.cloudinary_url(
        document.cloudinary_public_id,
        resource_type=resource_type,
        secure=True,
        flags="attachment"
    )

    return redirect(download_url)

@main.route(
    "/customer/<string:customer_code>/document/<int:document_id>/delete",
    methods=["POST"]
)

@login_required
def delete_customer_document(customer_code, document_id):

    # Make sure the customer belongs to the logged-in user
    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id
    ).first_or_404()

    # Make sure the document belongs to this customer
    document = CustomerDocument.query.filter_by(
        id=document_id,
        customer_id=customer.id
    ).first_or_404()

    # Delete the file from Cloudinary
    if document.cloudinary_public_id:
        configure_cloudinary()

        cloudinary.uploader.destroy(
            document.cloudinary_public_id,
            resource_type=document.cloudinary_resource_type or "image"
        )

    # Delete database record
    db.session.delete(document)
    db.session.commit()

    flash("Document deleted successfully.", "success")

    return redirect(
        url_for(
            "main.customer_details",
            customer_code=customer.customer_code
        )
    )    
    

@main.route("/archive_customer/<string:customer_code>")
@login_required
def archive_customer(customer_code):

    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id,
        is_archived=False
    ).first_or_404()

    customer.is_archived = True

    db.session.commit()

    flash("Customer archived successfully!", "success")

    return redirect(url_for("main.customers"))


@main.route("/archived_customers")
@login_required
def archived_customers():

    customers = Customer.query.filter_by(
        created_by=current_user.id,
        is_archived=True
    ).all()

    return render_template(
        "archived_customers.html",
        customers=customers
    )



@main.route("/restore_customer/<string:customer_code>")
@login_required
def restore_customer(customer_code):

    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id,
        is_archived=True
    ).first_or_404()

    customer.is_archived = False
    db.session.commit()

    flash("Customer restored successfully!", "success")

    return redirect(url_for("main.archived_customers"))



@main.route("/delete_customer/<string:customer_code>")
@login_required
def delete_customer(customer_code):

    customer = Customer.query.filter_by(
        customer_code=customer_code,
        created_by=current_user.id,
        is_archived=True
    ).first_or_404()

    db.session.delete(customer)
    db.session.commit()

    flash("Customer permanently deleted.", "success")

    return redirect(url_for("main.archived_customers"))


@main.route("/profile", methods=["GET", "POST"])
@login_required
def profile():

    form = ProfileForm()

    print("Request method:", request.method)

    if request.method == "POST":
        print("POST received")

    # Populate the form when opening the page
    if request.method == "GET":
        form.username.data = current_user.username
        form.email.data = current_user.email
        form.organisation_name.data = current_user.organisation_name
        form.business_phone.data = current_user.business_phone
        form.business_email.data = current_user.business_email
        form.business_address.data = current_user.business_address
        form.website.data = current_user.website

    if form.validate_on_submit():

        print("Form validated successfully!")

        current_user.username = form.username.data
        current_user.email = form.email.data
        current_user.organisation_name = form.organisation_name.data
        current_user.business_phone = form.business_phone.data
        current_user.business_email = form.business_email.data
        current_user.business_address = form.business_address.data
        current_user.website = form.website.data
        # Upload organisation logo to Cloudinary
        if form.organisation_logo.data:
            upload_result = cloudinary.uploader.upload(
                form.organisation_logo.data,
                folder=f"organisation_logos/{current_user.id}",
                resource_type="image"
            )

            current_user.organisation_logo = upload_result["secure_url"]

        try:
            db.session.commit()
            flash("Profile updated successfully!", "success")

        except Exception as e:
            db.session.rollback()
            print("Database error:", e)
            flash(f"Database error: {e}", "danger")

        return redirect(url_for("main.profile"))

    # Show validation errors (only if the form failed validation)
    if request.method == "POST" and form.errors:
        print("Validation errors:", form.errors)

    return render_template(
        "profile.html",
        form=form
    )
        
        
@main.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():

    form = ForgotPasswordForm()

    if form.validate_on_submit():

        email = form.email.data

        user = User.query.filter_by(email=email).first()

        if user:
            print("User found:", user.email)

            from app.email import send_reset_email

            print("About to send email...")

            send_reset_email(user)

            print("send_reset_email() finished")

        else:
            print("User NOT found")

        flash("If an account exists for that email address, a password reset link will be sent.")

        return redirect(url_for('main.login'))

    return render_template('forgot_password.html', form=form)
    
@main.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):

    user = User.verify_reset_token(token)

    if user is None:
        flash(
            "That password reset link is invalid or has expired.",
            "danger"
        )
        return redirect(url_for("main.forgot_password"))

    form = ResetPasswordForm()

    if form.validate_on_submit():

        user.password = generate_password_hash(
            form.password.data
        )

        try:
            db.session.commit()

            flash(
                "Your password has been updated successfully. Please log in.",
                "success"
            )

            return redirect(url_for("main.login"))

        except Exception as e:

            db.session.rollback()

            print("Database error:", e)

            flash(
                "An error occurred while updating your password.",
                "danger"
            )

    return render_template(
        "reset_password.html",
        form=form
    )

   

       

